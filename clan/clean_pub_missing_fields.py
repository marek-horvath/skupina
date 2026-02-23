#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Remove rows from an Excel file where 'authors' or 'venue' is empty.

Default:
  Input : pub.xlsx
  Output: pub.xlsx  (in-place update)

Worksheet default: Publications
"""

import argparse
from pathlib import Path
import openpyxl


def is_empty(value) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    return s == ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", default="pub.xlsx", help="Input XLSX file (default: pub.xlsx)")
    ap.add_argument("--sheet", dest="sheet", default="Publications", help="Worksheet name (default: Publications)")
    args = ap.parse_args()

    path = Path(args.inp)
    if not path.exists():
        raise SystemExit(f"File not found: {path.resolve()}")

    wb = openpyxl.load_workbook(path)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}")

    ws = wb[args.sheet]

    # Identify columns by header
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    if "authors" not in headers or "venue" not in headers:
        raise SystemExit("Required columns 'authors' and/or 'venue' not found.")

    authors_col = headers.index("authors") + 1
    venue_col = headers.index("venue") + 1

    rows_to_delete = []

    for r in range(2, ws.max_row + 1):
        authors = ws.cell(r, authors_col).value
        venue = ws.cell(r, venue_col).value

        if is_empty(authors) or is_empty(venue):
            rows_to_delete.append(r)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    wb.save(path)

    print(f"Removed rows: {len(rows_to_delete)}")
    print(f"File updated: {path.resolve()}")


if __name__ == "__main__":
    main()
