#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Remove duplicate rows from an Excel file based on the 'title' column.

Default:
  Input : pub.xlsx
  Output: pub.xlsx  (in-place update)

Keeps the first occurrence of each title and removes the rest.
"""

import argparse
from pathlib import Path
import openpyxl


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", default="pub.xlsx", help="Input XLSX file (default: pub.xlsx)")
    ap.add_argument("--sheet", dest="sheet", default="Publications", help="Worksheet name (default: Publications)")
    args = ap.parse_args()

    path = Path(args.inp)
    if not path.exists():
        raise SystemExit(f"File not found: {path.resolve()}")

    wb = openpyxl.load_workbook(path)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}")

    ws = wb[args.sheet]

    # Identify column indexes by header
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "title" not in headers:
        raise SystemExit("Column 'title' not found in header.")

    title_col = headers.index("title") + 1

    seen = set()
    rows_to_delete = []

    for r in range(2, ws.max_row + 1):
        title = str(ws.cell(r, title_col).value or "").strip().lower()
        if title in seen:
            rows_to_delete.append(r)
        else:
            seen.add(title)

    # Delete from bottom to top to keep indices valid
    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    wb.save(path)

    print(f"Removed duplicates: {len(rows_to_delete)}")
    print(f"File updated: {path.resolve()}")


if __name__ == "__main__":
    main()
