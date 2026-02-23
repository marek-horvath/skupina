#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import unicodedata
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


DEFAULT_INPUT = "exp.rtf"
DEFAULT_XLSX = "pub.xlsx"
SHEET_NAME = "pub"
HEADERS = ["date", "title", "authors", "venue", "link", "type"]

ENTRY_START_RE = re.compile(r"^\s*(ADC|ADD|ADM|V[123])\d{3}\s+\[\d+\]\s+(.+?)\s*$")
YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")
ACCESS_RE = re.compile(r"Spôsob prístupu:\s*([^\s]+)")
IN_RE = re.compile(r"\bIn:\s*(.+?)\s*\.")


def clean_text(s: str) -> str:
    """
    Normalizácia textu pre XLSX:
    - Unicode NFC (spojí rozložené diakritiky do bežných znakov)
    - zruší NBSP, BOM, zero-width
    - zjednotí „smart quotes“ na jednoduché úvodzovky
    - zjednotí pomlčky
    """
    if s is None:
        return ""

    # základné odstránenie problematických invisibles
    s = s.replace("\ufeff", "")  # BOM
    s = s.replace("\u200b", "")  # zero-width space
    s = s.replace("\u200c", "")
    s = s.replace("\u200d", "")
    s = s.replace("\u2060", "")  # word joiner
    s = s.replace("\xa0", " ")   # NBSP -> space

    # smart quotes/apostrofy -> ASCII
    s = (s.replace("’", "'")
           .replace("‘", "'")
           .replace("“", '"')
           .replace("”", '"')
           .replace("„", '"')
           .replace("‟", '"'))

    # pomlčky
    s = (s.replace("–", "-")
           .replace("—", "-")
           .replace("−", "-"))

    # Unicode normalizácia (diakritika do “normálnych” znakov)
    s = unicodedata.normalize("NFC", s)

    # whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


def rtf_to_text(s: str) -> str:
    """
    Minimalistické "odrtfovanie" exportu:
    - preloží \uNNNN? na Unicode
    - preloží \'hh (latin-1) na znak
    - odstráni RTF control words (\b0, \par, \fs..., atď.)
    - odstráni zoskupovacie { }
    """
    def _u_repl(m: re.Match) -> str:
        n = int(m.group(1))
        n = n % 65536
        return chr(n)

    s = re.sub(r"\\u(-?\d+)\??.", _u_repl, s)

    def _hex_repl(m: re.Match) -> str:
        return bytes([int(m.group(1), 16)]).decode("latin-1", errors="ignore")

    s = re.sub(r"\\'([0-9a-fA-F]{2})", _hex_repl, s)

    # niektoré symbolické escape sekvencie
    s = s.replace(r"\{", "{").replace(r"\}", "}")
    s = s.replace(r"\~", " ").replace(r"\_", " ")

    # odstráň control words
    s = re.sub(r"\\[a-zA-Z]+\d*\s?", "", s)

    # odstráň zvyšné backslash znaky
    s = s.replace("\\", "")

    # odstráň grouping braces
    s = s.replace("{", "").replace("}", "")

    return s


def category_to_type(cat: str) -> str:
    # len journal / conference; V1 sa mapuje ako journal
    return "conference" if cat == "V2" else "journal"


def extract_title_from_tail(tail: str) -> str:
    if " / " in tail:
        return clean_text(tail.split(" / ", 1)[0])
    return clean_text(tail)


def extract_authors(block_lines: List[str]) -> str:
    for ln in block_lines:
        ln = ln.strip()
        if ln.startswith("[") and "]" in ln:
            content = ln[1:ln.rfind("]")].strip()
            parts = [clean_text(p) for p in content.split(" - ") if clean_text(p)]
            if parts:
                return "; ".join(parts)
    return ""


def extract_venue(block_lines: List[str]) -> str:
    for ln in block_lines:
        m = IN_RE.search(ln)
        if m:
            return clean_text(m.group(1))
    return ""


def extract_link(block_lines: List[str]) -> str:
    for ln in block_lines:
        m = ACCESS_RE.search(ln)
        if m:
            link = clean_text(m.group(1)).rstrip(".")
            return link
    return ""


def extract_year(block_lines: List[str]) -> str:
    joined = " ".join(block_lines)
    m = YEAR_RE.search(joined)
    return m.group(0) if m else ""


def split_into_blocks(lines: List[str]) -> List[Tuple[str, List[str]]]:
    blocks: List[Tuple[str, List[str]]] = []
    current_cat: Optional[str] = None
    current_block: List[str] = []

    for raw in lines:
        line = raw.rstrip("\n")
        m = ENTRY_START_RE.match(line)
        if m:
            if current_cat is not None and current_block:
                blocks.append((current_cat, current_block))
            current_cat = m.group(1)
            current_block = [line]
        else:
            if current_cat is not None:
                current_block.append(line)

    if current_cat is not None and current_block:
        blocks.append((current_cat, current_block))

    return blocks


def parse_record(category: str, block_lines: List[str]) -> Dict[str, str]:
    start = block_lines[0].strip()
    m = ENTRY_START_RE.match(start)
    if not m:
        return {}

    cat = m.group(1)
    tail = m.group(2)

    title = extract_title_from_tail(tail)
    authors = extract_authors(block_lines)
    venue = extract_venue(block_lines)
    link = extract_link(block_lines)
    year = extract_year(block_lines)
    rec_type = category_to_type(cat)

    return {
        "date": clean_text(year),
        "title": clean_text(title),
        "authors": clean_text(authors),
        "venue": clean_text(venue),
        "link": clean_text(link),
        "type": clean_text(rec_type),
    }


def get_or_create_workbook(xlsx_path: str):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    return wb, ws


def ensure_header(ws) -> None:
    # ak je sheet prázdny, doplň hlavičku
    if ws.max_row == 0:
        ws.append(HEADERS)
        return
    # ak prvý riadok nie je hlavička, doplň ju navrch (konzervatívne: pridáme nový riadok 1)
    first = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
    first_norm = [clean_text(str(v)) if v is not None else "" for v in first]
    if first_norm != HEADERS:
        ws.insert_rows(1)
        for i, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=i, value=h)


def append_rows_to_xlsx(xlsx_path: str, rows: List[Dict[str, str]]) -> None:
    wb, ws = get_or_create_workbook(xlsx_path)
    ensure_header(ws)

    for r in rows:
        ws.append([r.get(h, "") for h in HEADERS])

    wb.save(xlsx_path)


def main() -> int:
    input_path = DEFAULT_INPUT
    output_path = DEFAULT_XLSX

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")

    raw = open(input_path, "r", encoding="utf-8", errors="ignore").read()
    text = rtf_to_text(raw)
    text = clean_text(text)  # celková normalizácia (NFC, quotes, whitespace)

    # potom rozbiť na riadky, ale zachovať zmysluplné delenie:
    # v RTF exportoch často "odrtfovanie" zlepuje; preto to ešte rozsekáme na pôvodné newline
    # ak by sa newline stratili, skript aj tak nájde ENTRY_START_RE v priebežných úsekoch len ťažko.
    # Preto čítame ešte raz s newline: použijeme raw->rtf_to_text bez clean_text zlepujúceho newline.
    text2 = rtf_to_text(raw)
    text2 = unicodedata.normalize("NFC", text2)
    # normalizuj problematické znaky, ale nestrácaj newline
    for bad, good in [
        ("\ufeff", ""), ("\u200b", ""), ("\u200c", ""), ("\u200d", ""), ("\u2060", ""),
        ("\xa0", " "), ("’", "'"), ("‘", "'"), ("“", '"'), ("”", '"'), ("„", '"'),
        ("–", "-"), ("—", "-"), ("−", "-"),
    ]:
        text2 = text2.replace(bad, good)

    lines = []
    for ln in text2.splitlines():
        ln = clean_text(ln)
        if ln:
            lines.append(ln)

    blocks = split_into_blocks(lines)

    rows: List[Dict[str, str]] = []
    for cat, block in blocks:
        rec = parse_record(cat, block)
        if rec and rec.get("title"):
            # final pass cleaning
            rec = {k: clean_text(v) for k, v in rec.items()}
            rows.append(rec)

    append_rows_to_xlsx(output_path, rows)
    print(f"OK: appended {len(rows)} rows to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
