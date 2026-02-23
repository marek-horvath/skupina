#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Convert EPC TUKE-style RTF publication export (default: exp.rtf) into a readable Excel file (default: pub.xlsx).

- Parses individual publication records (e.g., ADC001, V2004, V3003).
- Extracts: date (year), title, authors, venue, link, type
- Appends rows to an existing pub.xlsx (creates it if missing).
- Normalizes Slovak diacritics (NFC) and cleans typical RTF artifacts.

Usage:
  python rtf_export_to_pub_xlsx.py
  python rtf_export_to_pub_xlsx.py --in other_export.rtf --out pub.xlsx
  python rtf_export_to_pub_xlsx.py --sheet Publications --dedupe
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# -----------------------------
# RTF -> plain text
# -----------------------------
def rtf_to_text(rtf: str) -> str:
    """
    Best-effort RTF to text.
    Tries striprtf if available; otherwise uses a lightweight fallback.
    """
    try:
        from striprtf.striprtf import rtf_to_text as _rtf_to_text  # type: ignore
        return _rtf_to_text(rtf)
    except Exception:
        pass

    # Fallback: decode hex escapes \'hh first
    def _hex_unescape(m: re.Match) -> str:
        try:
            return bytes([int(m.group(1), 16)]).decode("cp1252", errors="replace")
        except Exception:
            return m.group(0)

    s = re.sub(r"\\'([0-9a-fA-F]{2})", _hex_unescape, rtf)

    # Remove some RTF groups
    s = re.sub(r"{\\\*[^{}]*}", "", s)
    s = re.sub(r"{\\fonttbl[^{}]*}", "", s)
    s = re.sub(r"{\\colortbl[^{}]*}", "", s)

    # Replace common control words with whitespace/newlines
    s = s.replace("\\par", "\n")
    s = s.replace("\\line", "\n")
    s = s.replace("\\tab", "\t")

    # Remove remaining control words like \b0 \fs24 \uXXXX?
    s = re.sub(r"\\u(-?\d+)\??", lambda m: chr(int(m.group(1)) % 65536), s)
    s = re.sub(r"\\[a-zA-Z]+-?\d*\s?", "", s)

    # Drop braces
    s = s.replace("{", "").replace("}", "")

    # Normalize whitespace
    s = re.sub(r"[ \t]+\n", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s


def norm(s: str) -> str:
    s = s.replace("\u00a0", " ").strip()
    s = unicodedata.normalize("NFC", s)
    s = re.sub(r"[ \t]{2,}", " ", s)
    return s


# -----------------------------
# Parsing EPC-like records
# -----------------------------
REC_START_RE = re.compile(r"^(?P<code>[A-Z]\d{3}\d{0,2}|[A-Z]{3}\d{3})\s+\[\d+\]\s+(?P<rest>.+)$")
# Example:
# V3003 [312156] Teach Programming ... Spôsob prístupu: http://dx.doi.org/... .
#
# Authors are often on a separate line:
# [PORUBÄN, Jaroslav - NOSÁL’, Milan - SULÍR, Matúš - CHODAREV, Sergej]


def _fix_apostrophes(s: str) -> str:
    return s.replace("’", "'").replace("‛", "'").replace("´", "'").replace("`", "'")


def _smart_title_token(token: str) -> str:
    """
    Title-case a token that may be all-caps and may contain hyphens or apostrophes.
    Keeps separators as-is.
    """
    token = _fix_apostrophes(token)
    parts = re.split(r"([\-'])", token.lower())
    out = []
    for p in parts:
        if p in {"-", "'"}:
            out.append(p)
        elif p:
            out.append(p[0].upper() + p[1:])
    return "".join(out)


def parse_authors_from_bracket_line(line: str) -> str:
    """
    Input line: [SURNAME, Name - SURNAME2, Name2 - ...]
    Output:     Name Surname; Name2 Surname2; ...
    """
    line = line.strip()
    if line.startswith("[") and line.endswith("]"):
        line = line[1:-1].strip()

    # Split on " - " (EPC export)
    parts = [p.strip() for p in line.split(" - ") if p.strip()]
    out = []
    for p in parts:
        p = _fix_apostrophes(p)
        if "," in p:
            surname, name = [x.strip() for x in p.split(",", 1)]
            surname = _smart_title_token(surname)
            out.append(f"{name} {surname}".strip())
        else:
            out.append(p)
    return "; ".join(out)


def extract_year(text: str) -> Optional[int]:
    m = re.search(r"\s-\s(19\d{2}|20\d{2})\b", text)
    if m:
        return int(m.group(1))
    m = re.search(r"\b(19\d{2}|20\d{2})\b", text)
    return int(m.group(1)) if m else None


def extract_title(rest: str) -> str:
    if " / " in rest:
        return rest.split(" / ", 1)[0].strip()
    return rest.strip()


def extract_venue(text: str) -> str:
    m = re.search(r"\bIn:\s*([^-\n\r]+?)(?:\s-\s|$)", text)
    if m:
        return m.group(1).strip().rstrip(".")
    return ""


def clean_url(url: str) -> str:
    url = _fix_apostrophes(url).strip()
    # Remove trailing punctuation that often appears in exports (especially a final dot).
    url = re.sub(r"[)\].,;:]+$", "", url)
    return url


def extract_link(text: str) -> str:
    # Common EPC export form: Spôsob prístupu: \* HYPERLINK "https://..."
    m = re.search(r'Spôsob prístupu:\s*\\\*\s*HYPERLINK\s+"([^"]+)"', text)
    if m:
        return clean_url(m.group(1))

    # Alternative: Spôsob prístupu: http://...
    m = re.search(r"Spôsob prístupu:\s*([^\s]+)", text)
    if m:
        return clean_url(m.group(1))

    # Fallback: first URL in text
    m = re.search(r"\b(https?://\S+)", text)
    if m:
        return clean_url(m.group(1))
    return ""



def infer_type(code: str) -> str:
    # Best-effort mapping (aligns with your CSV categories)
    prefix = re.match(r"^[A-Z]+", code).group(0) if re.match(r"^[A-Z]+", code) else code
    if prefix in {"ADC", "ADD", "ADM", "V3"}:
        return "journal"
    if prefix in {"V2"}:
        return "conference"
    if prefix in {"V1"}:
        return "book"
    return ""


def parse_records(plain: str) -> List[Dict[str, str]]:
    """
    Build records by:
    - finding record start lines (ADC001..., V2004..., etc.)
    - consuming subsequent lines until next record start
    - reading authors from the separate bracket line inside the block
    """
    lines = [norm(_fix_apostrophes(x)) for x in plain.splitlines()]
    recs: List[Dict[str, str]] = []

    i = 0
    while i < len(lines):
        line = lines[i]
        m = REC_START_RE.match(line)
        if not m:
            i += 1
            continue

        code = m.group("code")
        rest = m.group("rest")

        # collect continuation lines until next record start
        block_lines = [line]
        j = i + 1
        while j < len(lines):
            nxt = lines[j]
            if REC_START_RE.match(nxt):
                break
            block_lines.append(nxt)
            j += 1

        # authors: usually a dedicated line like "[SURNAME, Name - ...]"
        authors = ""
        for bl in reversed(block_lines):
            bls = bl.strip()
            if bls.startswith("[") and bls.endswith("]") and "," in bls and " - " in bls:
                authors = parse_authors_from_bracket_line(bls)
                break

        block_text = " ".join([b for b in block_lines if b]).strip()
        block_text = re.sub(r"\s{2,}", " ", block_text)

        year = extract_year(block_text)
        title = extract_title(rest)
        venue = extract_venue(block_text)
        link = extract_link(block_text)
        ptype = infer_type(code)

        recs.append({
            "date": str(year) if year else "",
            "title": norm(title),
            "authors": norm(authors),
            "venue": norm(venue),
            "link": norm(link),
            "type": ptype,
        })

        i = j

    return recs


# -----------------------------
# Excel output (append + format)
# -----------------------------
COLUMNS = ["date", "title", "authors", "venue", "link", "type"]


def load_existing_rows(wb_path: Path, sheet_name: str) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    if wb_path.exists():
        wb = openpyxl.load_workbook(wb_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        return wb, ws

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws


def ensure_header(ws) -> None:
    if ws.max_row == 1 and all((ws.cell(1, c).value is None for c in range(1, len(COLUMNS) + 1))):
        for c, name in enumerate(COLUMNS, start=1):
            ws.cell(row=1, column=c, value=name)

    header_font = Font(bold=True)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def get_existing_keys(ws, dedupe: bool) -> set:
    if not dedupe or ws.max_row < 2:
        return set()

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h): i + 1 for i, h in enumerate(header) if h}

    keys = set()
    for r in range(2, ws.max_row + 1):
        title = str(ws.cell(r, idx.get("title", 2)).value or "").strip()
        link = str(ws.cell(r, idx.get("link", 5)).value or "").strip()
        year = str(ws.cell(r, idx.get("date", 1)).value or "").strip()
        keys.add((link.lower(), title.lower(), year))
    return keys


def append_rows(ws, rows: List[Dict[str, str]], dedupe: bool) -> int:
    ensure_header(ws)
    existing = get_existing_keys(ws, dedupe)

    appended = 0
    for rec in rows:
        title = (rec.get("title") or "").strip()
        link = (rec.get("link") or "").strip()
        year = (rec.get("date") or "").strip()
        key = (link.lower(), title.lower(), year)

        if dedupe and key in existing:
            continue

        ws.append([rec.get(col, "") for col in COLUMNS])
        appended += 1
        if dedupe:
            existing.add(key)

    return appended


def autosize_columns(ws, max_width: int = 80) -> None:
    for col_idx in range(1, len(COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), max_width)

    wrap_cols = {"title", "authors", "venue", "link"}
    for col_idx, name in enumerate(COLUMNS, start=1):
        if name in wrap_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", default="exp.rtf", help="Input RTF file (default: exp.rtf)")
    ap.add_argument("--out", dest="outp", default="pub.xlsx", help="Output XLSX file (default: pub.xlsx)")
    ap.add_argument("--sheet", dest="sheet", default="Publications", help="Worksheet name (default: Publications)")
    ap.add_argument("--dedupe", action="store_true", help="Skip duplicates based on (link,title,year)")
    args = ap.parse_args()

    in_path = Path(args.inp)
    out_path = Path(args.outp)

    if not in_path.exists():
        raise SystemExit(f"Input file not found: {in_path.resolve()}")

    raw = in_path.read_text(encoding="utf-8", errors="replace")
    plain = unicodedata.normalize("NFC", rtf_to_text(raw))

    records = parse_records(plain)
    if not records:
        raise SystemExit("No publication records detected. The input may not match the expected EPC export structure.")

    wb, ws = load_existing_rows(out_path, args.sheet)
    appended = append_rows(ws, records, dedupe=args.dedupe)
    autosize_columns(ws)
    wb.save(out_path)

    print(f"Parsed records: {len(records)}")
    print(f"Appended rows:  {appended}")
    print(f"Output:         {out_path.resolve()}")


if __name__ == "__main__":
    main()
